import csv
import json
import os
import re
import uuid
from pathlib import Path

from django.conf import settings
from django.http import FileResponse, Http404, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from .upload_handlers import StreamingFileUploadHandler

UPLOAD_DIR = Path(settings.BASE_DIR) / 'uploads'
FILE_ID_RE = re.compile(r'^[0-9a-f]{32}\.(csv|xls|xlsx|xlsm|xlsb)$')
MAX_LIMIT = 2000
SAMPLE_COUNT = 10


@csrf_exempt
@require_http_methods(["POST"])
def upload(request):
    handler = StreamingFileUploadHandler(request)
    request.upload_handlers = [handler]

    uploaded = request.FILES.get('file')

    if handler.rejected:
        return JsonResponse({'error': handler.reject_reason}, status=400)

    if uploaded is None:
        return JsonResponse({'error': 'no file provided'}, status=400)

    return JsonResponse({
        'id': getattr(uploaded, 'file_id', None),
        'name': uploaded.name,
        'size': uploaded.size,
        'content_type': uploaded.content_type,
    })


def _read_csv_page(path, offset, limit):
    with open(path, 'r', newline='', encoding='utf-8-sig', errors='replace') as f:
        reader = csv.reader(f)
        try:
            columns = next(reader)
        except StopIteration:
            return [], [], False

        for _ in range(offset):
            try:
                next(reader)
            except StopIteration:
                return columns, [], False

        rows = []
        for _ in range(limit):
            try:
                rows.append(next(reader))
            except StopIteration:
                return columns, rows, False

        try:
            next(reader)
            has_more = True
        except StopIteration:
            has_more = False
        return columns, rows, has_more


def _stringify(cell):
    if cell is None:
        return ''
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def _read_xlsx_page(path, offset, limit):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return [], [], False

        columns = [_stringify(c) for c in header]

        for _ in range(offset):
            try:
                next(rows_iter)
            except StopIteration:
                return columns, [], False

        rows = []
        for _ in range(limit):
            try:
                r = next(rows_iter)
                rows.append([_stringify(c) for c in r])
            except StopIteration:
                return columns, rows, False

        try:
            next(rows_iter)
            has_more = True
        except StopIteration:
            has_more = False
        return columns, rows, has_more
    finally:
        wb.close()


def _read_xls_page(path, offset, limit):
    import xlrd

    book = xlrd.open_workbook(str(path), on_demand=True)
    try:
        sheet = book.sheet_by_index(0)
        nrows = sheet.nrows
        if nrows == 0:
            return [], [], False

        columns = [_stringify(c) for c in sheet.row_values(0)]
        start = 1 + offset
        end = min(start + limit, nrows)
        rows = [[_stringify(c) for c in sheet.row_values(i)] for i in range(start, end)]
        has_more = end < nrows
        return columns, rows, has_more
    finally:
        book.release_resources()


def _read_xlsb_page(path, offset, limit):
    from pyxlsb import open_workbook

    with open_workbook(str(path)) as wb:
        sheet_name = wb.sheets[0]
        with wb.get_sheet(sheet_name) as sheet:
            rows_iter = sheet.rows()
            try:
                header = next(rows_iter)
            except StopIteration:
                return [], [], False

            columns = [_stringify(c.v) for c in header]

            for _ in range(offset):
                try:
                    next(rows_iter)
                except StopIteration:
                    return columns, [], False

            rows = []
            for _ in range(limit):
                try:
                    r = next(rows_iter)
                    rows.append([_stringify(c.v) for c in r])
                except StopIteration:
                    return columns, rows, False

            try:
                next(rows_iter)
                has_more = True
            except StopIteration:
                has_more = False
            return columns, rows, has_more


@require_http_methods(["GET"])
def preview(request, file_id):
    if not FILE_ID_RE.match(file_id):
        return JsonResponse({'error': 'invalid id'}, status=400)

    path = UPLOAD_DIR / file_id
    if not path.exists():
        return JsonResponse({'error': 'not found'}, status=404)

    try:
        offset = max(0, int(request.GET.get('offset', '0')))
        limit = min(MAX_LIMIT, max(1, int(request.GET.get('limit', '100'))))
    except ValueError:
        return JsonResponse({'error': 'bad offset/limit'}, status=400)

    ext = path.suffix.lower()
    try:
        if ext == '.csv':
            columns, rows, has_more = _read_csv_page(path, offset, limit)
        elif ext in {'.xlsx', '.xlsm'}:
            columns, rows, has_more = _read_xlsx_page(path, offset, limit)
        elif ext == '.xls':
            columns, rows, has_more = _read_xls_page(path, offset, limit)
        elif ext == '.xlsb':
            columns, rows, has_more = _read_xlsb_page(path, offset, limit)
        else:
            return JsonResponse(
                {'error': f'preview not supported for {ext} files'}, status=415
            )
    except Exception as e:
        return JsonResponse({'error': f'failed to parse file: {e}'}, status=500)

    return JsonResponse({
        'columns': columns or [],
        'rows': rows,
        'offset': offset,
        'limit': limit,
        'has_more': has_more,
    })


def _read_all(path):
    ext = path.suffix.lower()
    if ext == '.csv':
        with open(path, 'r', newline='', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.reader(f)
            try:
                columns = next(reader)
            except StopIteration:
                return [], []
            rows = [list(r) for r in reader]
            return columns, rows
    if ext in {'.xlsx', '.xlsm'}:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                return [], []
            columns = [_stringify(c) for c in header]
            rows = [[_stringify(c) for c in r] for r in rows_iter]
            return columns, rows
        finally:
            wb.close()
    if ext == '.xls':
        import xlrd
        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheet = book.sheet_by_index(0)
            if sheet.nrows == 0:
                return [], []
            columns = [_stringify(c) for c in sheet.row_values(0)]
            rows = [[_stringify(c) for c in sheet.row_values(i)] for i in range(1, sheet.nrows)]
            return columns, rows
        finally:
            book.release_resources()
    if ext == '.xlsb':
        from pyxlsb import open_workbook
        with open_workbook(str(path)) as wb:
            with wb.get_sheet(wb.sheets[0]) as sheet:
                rows_iter = sheet.rows()
                try:
                    header = next(rows_iter)
                except StopIteration:
                    return [], []
                columns = [_stringify(c.v) for c in header]
                rows = [[_stringify(c.v) for c in r] for r in rows_iter]
                return columns, rows
    raise ValueError(f'unsupported format {ext}')


def _column_samples(path, target_column, n=SAMPLE_COUNT):
    columns, rows = _read_all(path)
    if target_column not in columns:
        return []
    idx = columns.index(target_column)
    samples = []
    for r in rows:
        if idx < len(r):
            val = r[idx]
            if val:
                samples.append(val)
        if len(samples) >= n:
            break
    return samples


@csrf_exempt
@require_http_methods(["POST"])
def generate_regex(request):
    try:
        data = json.loads(request.body or b'{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'invalid JSON'}, status=400)

    file_id = data.get('file_id', '')
    target_column = data.get('target_column', '') or ''
    description = (data.get('description') or '').strip()

    if not description:
        return JsonResponse({'error': 'description is required'}, status=400)
    if not FILE_ID_RE.match(file_id):
        return JsonResponse({'error': 'invalid file_id'}, status=400)

    path = UPLOAD_DIR / file_id
    if not path.exists():
        return JsonResponse({'error': 'file not found'}, status=404)

    if not os.environ.get('ANTHROPIC_API_KEY'):
        return JsonResponse(
            {'error': 'ANTHROPIC_API_KEY is not configured on the server'},
            status=500,
        )

    try:
        samples = _column_samples(path, target_column)
    except Exception as e:
        return JsonResponse({'error': f'failed to read file: {e}'}, status=500)

    sample_text = '\n'.join(f'- {s!r}' for s in samples) if samples else '(no samples)'

    from anthropic import Anthropic, APIError

    client = Anthropic()
    try:
        response = client.messages.create(
            model='claude-haiku-4-5',
            max_tokens=2048,
            output_config={
                'format': {
                    'type': 'json_schema',
                    'schema': {
                        'type': 'object',
                        'properties': {
                            'regex': {
                                'type': 'string',
                                'description': 'A Python-compatible regex pattern.',
                            },
                            'explanation': {
                                'type': 'string',
                                'description': 'One short sentence describing what the pattern matches.',
                            },
                            'replacement': {
                                'type': 'string',
                                'description': (
                                    'A sensible default replacement string for what '
                                    'the pattern matches. May be empty to delete '
                                    'matches. Use shape-preserving masks for things '
                                    'like phone numbers (***-***-****), generic '
                                    'placeholders for identities (REDACTED), and '
                                    'descriptive tags for content (e.g. [link removed]).'
                                ),
                            },
                        },
                        'required': ['regex', 'explanation', 'replacement'],
                        'additionalProperties': False,
                    },
                },
            },
            system=(
                "You convert natural-language descriptions of text patterns into "
                "Python-compatible regular expressions AND propose a sensible default "
                "replacement string for masking, redacting, or transforming each match. "
                "Output the regex as a plain string with no surrounding slashes, quotes, "
                "or flags. Prefer simple, robust patterns that handle common variations "
                "in the sample values. For the replacement, infer intent from the "
                "description (e.g. 'redact' -> REDACTED, 'mask phone' -> ***-***-****, "
                "'remove urls' -> [link removed]). If unsure, return an empty string."
            ),
            messages=[{
                'role': 'user',
                'content': (
                    f'Description: {description}\n'
                    f'Target column: {target_column}\n'
                    f'Sample values from the column:\n{sample_text}'
                ),
            }],
        )
    except APIError as e:
        return JsonResponse({'error': f'LLM error: {e}'}, status=502)
    except Exception as e:
        return JsonResponse({'error': f'unexpected error calling LLM: {e}'}, status=500)

    payload = None
    for block in response.content:
        if getattr(block, 'type', None) == 'text':
            try:
                payload = json.loads(block.text)
                break
            except json.JSONDecodeError:
                continue

    if payload is None:
        return JsonResponse({'error': 'no usable JSON in LLM response'}, status=500)

    regex_pat = (payload.get('regex') or '').strip()
    explanation = (payload.get('explanation') or '').strip()
    replacement = payload.get('replacement') or ''

    try:
        re.compile(regex_pat)
    except re.error as e:
        return JsonResponse({'error': f'LLM produced invalid regex: {e}'}, status=500)

    return JsonResponse({
        'regex': regex_pat,
        'explanation': explanation,
        'replacement': replacement,
    })


@csrf_exempt
@require_http_methods(["POST"])
def transform(request):
    try:
        data = json.loads(request.body or b'{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'invalid JSON'}, status=400)

    file_id = data.get('file_id', '')
    target_column = data.get('target_column', '') or ''
    regex_pat = data.get('regex', '') or ''
    replacement = data.get('replacement', '') or ''

    if not FILE_ID_RE.match(file_id):
        return JsonResponse({'error': 'invalid file_id'}, status=400)

    path = UPLOAD_DIR / file_id
    if not path.exists():
        return JsonResponse({'error': 'file not found'}, status=404)

    try:
        pattern = re.compile(regex_pat)
    except re.error as e:
        return JsonResponse({'error': f'invalid regex: {e}'}, status=400)

    try:
        columns, rows = _read_all(path)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=415)
    except Exception as e:
        return JsonResponse({'error': f'failed to read file: {e}'}, status=500)

    if target_column not in columns:
        return JsonResponse(
            {'error': f'column {target_column!r} not found'}, status=400
        )

    col_idx = columns.index(target_column)
    match_count = 0
    for row in rows:
        if col_idx >= len(row):
            continue
        original = row[col_idx] if row[col_idx] is not None else ''
        new_val, n = pattern.subn(replacement, str(original))
        row[col_idx] = new_val
        match_count += n

    download_id = f'{uuid.uuid4().hex}.csv'
    download_path = UPLOAD_DIR / download_id
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with open(download_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(rows)

    return JsonResponse({
        'download_id': download_id,
        'match_count': match_count,
        'row_count': len(rows),
    })


@require_http_methods(["GET"])
def download(request, file_id):
    if not FILE_ID_RE.match(file_id):
        return JsonResponse({'error': 'invalid id'}, status=400)
    path = UPLOAD_DIR / file_id
    if not path.exists():
        raise Http404
    return FileResponse(
        open(path, 'rb'),
        as_attachment=True,
        filename=f'transformed-{file_id}',
    )
